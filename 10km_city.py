from openpyxl import load_workbook
import asyncio
import time
from selenium import webdriver
from fake_useragent import UserAgent
from selenium.webdriver.common.keys import Keys
from datetime import datetime, timedelta


def city_tula():
    async def tula1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=26-Тула&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Оборонная улица, 83")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Оборонная 23", "Менделеевская 9", "Красноармейский проспект 8", "Красноармейский проспект 34",
                              "Луначарского, 63", "Пороховая 177", "Октябрьская улица, 217", "Октябрьская улица, 300", "Фестивальная улица, 10",
                                "13-й Горельский проезд 17"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тула"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Максим")
                driver.quit()
                continue
            driver.quit()
            break
    async def tula2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
#####Яндекс Тула
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Тула Оборонная улица, 83")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Оборонная 23", "Менделеевская 9", "Красноармейский проспект 8", "Красноармейский проспект 34",
                              "Луначарского, 63", "Пороховая 177", "Октябрьская улица, 217", "Октябрьская улица, 300", "Фестивальная улица, 10",
                                "13-й Горельский проезд 17"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok #[0],spisok[1][0:3],spisok[2][0:3],spisok[3][0:3]
                print(*a)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тула"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
    async def tula3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                #Омега сайт код 1
                driver.get("https://taxiomega.ru/?town=tula")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Оборонная улица, 83")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Оборонная 23", "Менделеевская 9", "Красноармейский проспект 8", "Красноармейский проспект 34",
                              "Луначарского, 63", "Пороховая 177", "Октябрьская улица, 217", "Октябрьская улица, 300", "Фестивальная улица, 10",
                                "13-й Горельский проезд 17"]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                if b[1] == "":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тула"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Омега")
                driver.quit()
                continue
            driver.quit()
            break
    async def tula4():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                #Поехали такси Тула1
                driver.get("https://taxipoehali.ru/client/?city=649-Тула&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Оборонная улица, 83")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Оборонная 23", "Менделеевская 9", "Красноармейский проспект 8", "Красноармейский проспект 34",
                              "Луначарского, 63", "Пороховая 177", "Октябрьская улица, 217", "Октябрьская улица, 300", "Фестивальная улица, 10",
                                "13-й Горельский проезд 17"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    time.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тула"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Поехали")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            tula1(),
            tula2(),
            tula3(),
            tula4()

        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())
    #код для КАЗАНИ ############
def city_kazan():
    async def kazan1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
        ##Omega
                driver.get("https://taxiomega.ru/?town=kazan")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Маршала Чуйкова 93")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Чуйкова 63а","Маршала Чуйкова, 37","Маршала Чуйкова, 11","Восстания, 14", "Восстания 40",
                           "Восстания, 80","Восстания, 110","Мамадышский тракт, 10Б",
                            "Академика Глушко, 47","Глушко 15" ]
                spisok4 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok4
                if [1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Казань"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Омега")
                driver.quit()
                continue
            driver.quit()
            break

#####Яндекс Казань
    async def kazan2():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Казань Маршала Чуйкова 93")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Чуйкова 63а","Маршала Чуйкова, 37","Маршала Чуйкова, 11","Восстания, 14", "Восстания 40",
                           "Восстания, 80","Восстания, 110","Мамадышский тракт, 10Б",
                            "Академика Глушко, 47","Глушко 15" ]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Казань"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекс")
                driver.quit()
                continue
            driver.quit()
            break
###Максим Казань2
    async def kazan3():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=19-Казань&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Маршала Чуйкова 93")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Чуйкова 63а","Маршала Чуйкова, 37","Маршала Чуйкова, 11","Восстания, 14", "Восстания 40",
                           "Восстания, 80","Восстания, 110","Мамадышский тракт, 10Б",
                            "Академика Глушко, 47","Глушко 15" ]
                spisok5 = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    time.sleep(3)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    driver.implicitly_wait(5)
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok5.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    driver.implicitly_wait(5)
                    i += 1
                d = spisok5
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Казань"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Максим")
                driver.quit()
                continue
            driver.quit()
            break

###Поехали Казань2
    async def kazan4():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=602-Казань&intl=ru-RU&referrer=self")
                driver.implicitly_wait(5)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Маршала Чуйкова 93")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Чуйкова 63а","Маршала Чуйкова, 37","Маршала Чуйкова, 11","Восстания, 14", "Восстания 40",
                           "Восстания, 80","Восстания, 110","Мамадышский тракт, 10Б",
                            "Академика Глушко, 47","Глушко 15" ]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    time.sleep(5)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    driver.implicitly_wait(5)
                    i += 1
                d = spisok2
                if d[1]== "":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Казань"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Поехали")
                driver.quit()
                continue
            driver.quit()
            break

####Микс Казань2
    async def kazan5():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://таксимикс.рф/city/kazan.html")
                driver.implicitly_wait(5)
                driver.get("http://таксимикс.рф/city/kazan.html")
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Маршала Чуйкова 93")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Чуйкова 63а","Маршала Чуйкова, 37","Маршала Чуйкова, 11","Восстания, 14", "Восстания 40",
                           "Восстания, 80","Восстания, 110","Мамадышский тракт, 10Б",
                            "Академика Глушко, 47","Глушко 15" ]
                spisok2 = ["Микс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Казань"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Микс")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            kazan1(),
            kazan2(),
            kazan3(),
            kazan4(),
            kazan5()

        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

        # код для третьего города
def city_ulan():
    async def ulan_ude1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)

    ######Омега УУ
                driver.get("https://taxiomega.ru/?town=ulan-ude")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Сахьяновой 9/1")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Ключевская 4д","Бийская 21","Красногвардейская улица, 33","Производственная улица, 6",
                            "Победы проспект, 7","Шмидта улица, 32","Партизанская 31","Буйко19",
                            "Коллективная улица, 13","Чертенкова 149"
                            ]
                spisok4 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    driver.implicitly_wait(5)
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    driver.implicitly_wait(5)
                    i += 1
                d = spisok4
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Улан"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Омега")
                driver.quit()
                continue
            driver.quit()
            break
####Максим УУУ
    async def ulan_ude2():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=54-Улан-Удэ&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Сахьяновой 9/1")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Ключевская 4д","Бийская 21","Красногвардейская улица, 33","Производственная улица, 6",
                            "Победы проспект, 7","Шмидта улица, 32","Партизанская 31","Буйко19",
                            "Коллективная улица, 13","Чертенкова 149"
                            ]
                spisok4 = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok4
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Улан"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Маским")
                driver.quit()
                continue
            driver.quit()
            break

                ####Поехали УУ
    async def ulan_ude3():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=660-Улан-Удэ&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Сахьяновой 9/1")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Ключевская 4д","Бийская 21","Красногвардейская улица, 33","Производственная улица, 6",
                            "Победы проспект, 7","Шмидта улица, 32","Партизанская 31","Буйко19",
                            "Коллективная улица, 13","Чертенкова 149"
                            ]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Улан"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Поехали")
                driver.quit()
                continue
            driver.quit()
            break

                ####Микс УУ
    async def ulan_ude4():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://таксимикс.рф/city/ulanude.html")
                driver.implicitly_wait(5)
                driver.get("http://таксимикс.рф/city/ulanude.html")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Сахьяновой 9/1")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Ключевская 4д","Бийская 21","Красногвардейская улица, 33","Производственная улица, 6",
                            "Победы проспект, 7","Шмидта улица, 32","Партизанская 31","Буйко19",
                            "Коллективная улица, 13","Чертенкова 149"
                            ]
                spisok2 = ["Микс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    driver.implicitly_wait(5)
                    i += 1
                d = spisok2
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Улан"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Микс")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Улан-Удэ
    async def ulan_ude5():
        while True:
            try:
                useragent = UserAgent()
                    #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Улан-Удэ Сахьяновой 9/1")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Ключевская 4д","Бийская 21","Красногвардейская улица, 33","Производственная улица, 6",
                            "Победы проспект, 7","Шмидта улица, 32","Партизанская 31","Буйко19",
                            "Коллективная улица, 13","Чертенкова 149"
                            ]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Улан"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            ulan_ude1(),
            ulan_ude2(),
            ulan_ude3(),
            ulan_ude4(),
            ulan_ude5()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

#ТОМСК
def city_tomsk():
    async def tomsk1():
        while True:
            try:
                print("Начинаю парсить данные по городу Томск")
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=13-Томск&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Мокрушина улица, 26")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Красноармейская улица, 152", "Нахимова улица, 9", "Красноармейская улица, 94",
                           "Красноармейская улица, 42","Гагарина улица, 38", "Яковлева улица, 41", "Дальне-Ключевская улица, 36",
                           "Дальне-Ключевская улица, 1", "Ленина проспект, 206","Ленина проспект, 242"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(3)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(3)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(3)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(3)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Томск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Максим")
                driver.quit()
                continue
            driver.quit()
            break

                #Омега сайт код Томск
    async def tomsk2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=tomsk")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Мокрушина улица, 26")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Красноармейская улица, 152", "Нахимова улица, 9", "Красноармейская улица, 94",
                           "Красноармейская улица, 42","Гагарина улица, 38", "Яковлева улица, 41", "Дальне-Ключевская улица, 36",
                           "Дальне-Ключевская улица, 1", "Ленина проспект, 206","Ленина проспект, 242"]
                spisok1 = ["Омега:   "]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Томск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Омега")
                driver.quit()
                continue
            driver.quit()
            break

#Поехали такси Томск
    async def tomsk3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=696-Томск&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(2)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Мокрушина улица, 26")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Красноармейская улица, 152", "Нахимова улица, 9", "Красноармейская улица, 94",
                           "Красноармейская улица, 42","Гагарина улица, 38", "Яковлева улица, 41", "Дальне-Ключевская улица, 36",
                           "Дальне-Ключевская улица, 1", "Ленина проспект, 206","Ленина проспект, 242"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(3)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Томск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Поехали")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Томск
    async def tomsk4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Томск Мокрушина улица, 26")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Красноармейская улица, 152", "Нахимова улица, 9", "Красноармейская улица, 94",
                           "Красноармейская улица, 42","Гагарина улица, 38", "Яковлева улица, 41", "Дальне-Ключевская улица, 36",
                           "Дальне-Ключевская улица, 1", "Ленина проспект, 206","Ленина проспект, 242"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Томск"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            tomsk1(),
            tomsk2(),
            tomsk3(),
            tomsk4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_tymen():
    async def tymen1():
        while True:
            try:
                print("Начинаю собирать данные по городу Тюмень")
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                ### Омега Тюмень
                driver.get("https://taxiomega.ru/?town=tumen")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("​Широтная 209")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Широтная 181","Широтная, 303", "Народная 63", "Широтная 79",
                            "Широтная, 35","30 лет Победы 31", "30 лет Победы 10","Мельникайте 101","Мельничная 49",
                            "Советская 88"]
                spisok4 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok4
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now()
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тюмень"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Омега")
                driver.quit()
                continue
            driver.quit()
            break
###Максим Тюмень
    async def tymen2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=3-Тюмень&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("​Широтная 209")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                await asyncio.sleep(4)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Широтная 181","Широтная, 303", "Народная 63", "Широтная 79",
                            "Широтная, 35","30 лет Победы 31", "30 лет Победы 10","Мельникайте 101","Мельничная 49",
                            "Советская 88"]
                spisok5 = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok5.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok5
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now()
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тюмень"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Максим")
                driver.quit()
                continue
            driver.quit()
            break

                ###Поехали Тюмень
    async def tymen3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=609-Тюмень&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("​Широтная 209")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                await asyncio.sleep(4)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Широтная 181","Широтная, 303", "Народная 63", "Широтная 79",
                            "Широтная, 35","30 лет Победы 31", "30 лет Победы 10","Мельникайте 101","Мельничная 49",
                            "Советская 88"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                print(*d)
                t = datetime.now()
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тюмень"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Поехали")
                driver.quit()
                continue
            driver.quit()
            break

                ####Микс Тюмень
    async def tymen4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://таксимикс.рф/city/tyumen.html")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Широтная 209")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                await asyncio.sleep(4)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Широтная 181","Широтная, 303", "Народная 63", "Широтная 79",
                            "Широтная, 35","30 лет Победы 31", "30 лет Победы 10","Мельникайте 101","Мельничная 49",
                            "Советская 88"]
                spisok2 = ["Микс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now()
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тюмень"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Микс")
                driver.quit()
                continue
            driver.quit()
            break

#####Яндекс Тюмень
    async def tymen5():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Тюмень Широтная 209")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Широтная 181","Широтная, 303", "Народная 63", "Широтная 79",
                            "Широтная, 35","30 лет Победы 31", "30 лет Победы 10","Мельникайте 101","Мельничная 49",
                            "Советская 88"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now()
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Тюмень"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            tymen1(),
            tymen2(),
            tymen3(),
            tymen4(),
            tymen5()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

        # код для третьего города
def city_kemerovo():
    async def kemerovo1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                ######Омега Кемерово
                print("Начинаю сбор данных в г. Кемерово")
                driver.get("https://taxiomega.ru/?town=kemerovo")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("​Тухачевского, 64")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Тухачевского, 56","Тухачевского, 40б","Тухачевского, 41А","Тухачевского, 16",
                            "Октябрьский проспект, 56","Пионерский бульвар, 13","Ленина проспект, 141",
                            "Ленина проспект, 39","Ленина проспект, 22",
                            "Якимова, 74"]
                spisok4 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok4
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Кемерово"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Омега")
                driver.quit()
                continue
            driver.quit()
            break
                ####Максим Кемерово
    async def kemerovo2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=12-Кемерово&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Тухачевского, 64")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Тухачевского, 56","Тухачевского, 40б","Тухачевского, 41А","Тухачевского, 16",
                            "Октябрьский проспект, 56","Пионерский бульвар, 13","Ленина проспект, 141",
                            "Ленина проспект, 39","Ленина проспект, 22",
                            "Якимова, 74"]
                spisok4 = ["Максим:"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    driver.implicitly_wait(5)
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok4
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Кемерово"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Максим")
                driver.quit()
                continue
            driver.quit()
            break

                ####Поехали Кемерово
    async def kemerovo3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=658-Кемерово&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(2)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Тухачевского, 64")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Тухачевского, 56","Тухачевского, 40б","Тухачевского, 41А","Тухачевского, 16",
                            "Октябрьский проспект, 56","Пионерский бульвар, 13","Ленина проспект, 141",
                            "Ленина проспект, 39","Ленина проспект, 22",
                            "Якимова, 74"]
                spisok2 = ["Поехали:"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    driver.implicitly_wait(5)
                    i += 1
                d = spisok2
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Кемерово"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Поехали")
                driver.quit()
                continue
            driver.quit()
            break

                ####Микс Кемерово
    async def kemerovo4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://таксимикс.рф/city/kemerovo.html")
                await asyncio.sleep(4)
                driver.get("http://таксимикс.рф/city/kemerovo.html")
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Тухачевского, 64")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Тухачевского, 56","Тухачевского, 40б","Тухачевского, 41А","Тухачевского, 16",
                            "Октябрьский проспект, 56","Пионерский бульвар, 13","Ленина проспект, 141",
                            "Ленина проспект, 39","Ленина проспект, 22",
                            "Якимова, 74"]
                spisok2 = ["Микс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Кемерово"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Микс")
                driver.quit()
                continue
            driver.quit()
            break

#####Яндекс Кемерово
    async def kemerovo5():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Кемерово Тухачевского, 64")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Тухачевского, 56","Тухачевского, 40б","Тухачевского, 41А","Тухачевского, 16",
                            "Октябрьский проспект, 56","Пионерский бульвар, 13","Ленина проспект, 141",
                            "Ленина проспект, 39","Ленина проспект, 22",
                            "Якимова, 74"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Кемерово"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            kemerovo1(),
            kemerovo2(),
            kemerovo3(),
            kemerovo4(),
            kemerovo5()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_surgut():
    async def surgut1():
        while True:
            try:

                print("Начинаю сбор данных в г.Сургут")
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                  #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=56-Сургут&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Привокзальная, 11")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Привокзальная 10","Привокзальная 28","Семена Билецкого 1","Игоря Киртбая 17", "Индустриальная 21",
                           "Индустриальная, 31","Производственная, 16","Нефтеюганское шоссе, 38",
                           "Университетская, 9","Университетская, 21" ]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽","")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                print(*b)
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Сургут"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Максим")
                driver.quit()
                continue
            driver.quit()
            break
    async def surgut2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=surgut")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Привокзальная, 11")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Привокзальная 10","Привокзальная 28","Семена Билецкого 1","Игоря Киртбая 17", "Индустриальная 21",
                           "Индустриальная, 31","Производственная, 16","Нефтеюганское шоссе, 38",
                           "Университетская, 9","Университетская, 21" ]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(3)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    driver.implicitly_wait(5)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    driver.implicitly_wait(5)
                    i += 1
                b = spisok1
                print(*b)
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Сургут"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Омега")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Сургут
    async def surgut3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Сургут Привокзальная, 11")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Привокзальная 10","Привокзальная 28","Семена Билецкого 1","Игоря Киртбая 17", "Индустриальная 21",
                           "Индустриальная, 31","Производственная, 16","Нефтеюганское шоссе, 38",
                           "Университетская, 9","Университетская, 21" ]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Сургут"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break

                #Поехали такси Сургут
    async def surgut4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=823-Сургут&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(2)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Привокзальная, 11")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Привокзальная 10","Привокзальная 28","Семена Билецкого 1","Игоря Киртбая 17", "Индустриальная 21",
                           "Индустриальная, 31","Производственная, 16","Нефтеюганское шоссе, 38",
                           "Университетская, 9","Университетская, 21" ]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                print(*b)
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Сургут"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Поехали")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            surgut1(),
            surgut2(),
            surgut3(),
            surgut4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())
            
def city_habarovsk():

    async def habarovsk1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                #####Яндекс Хабаровск
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Хабаровск Суворова25")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Суворова8", "Краснореченская70б", "Иртышская15", "Краснореченская 151",
                          "Краснореченская 100", "Краснореченская104", "Краснореченская207", "Юности30", "Пионерская1Е",
                            "Волочаевская113"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Хабаровск"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
    async def habarovsk2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=10-Хабаровск&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Суворова25")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Суворова8", "Краснореченская70б", "Иртышская15", "Краснореченская 151",
                          "Краснореченская 100", "Краснореченская104", "Краснореченская207", "Юности30", "Пионерская1Е",
                            "Волочаевская113"]
                spisok = ["Максим:"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                print(*b)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Хабаровск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка вселенной, начинаю проверять Максим")
                driver.quit()
                continue
            driver.quit()
            break

    async def habarovsk3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                #Омега сайт код 1
                driver.get("https://taxiomega.ru/?town=habarovsk")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Суворова25")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Суворова8", "Краснореченская70б", "Иртышская15", "Краснореченская 151",
                          "Краснореченская 100", "Краснореченская104", "Краснореченская207", "Юности30", "Пионерская1Е",
                            "Волочаевская113"]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                print(*b)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Хабаровск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка вселенной, начинаю проверять Омега")
                driver.quit()
                continue
            driver.quit()
            break
    async def habarovsk4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                #Поехали такси Хабаровск
                driver.get("https://taxipoehali.ru/client/?city=611-Хабаровск&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(2)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Суворова25")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Суворова8", "Краснореченская70б", "Иртышская15", "Краснореченская 151",
                          "Краснореченская 100", "Краснореченская104", "Краснореченская207", "Юности30", "Пионерская1Е",
                            "Волочаевская113"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                print(*b)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Хабаровск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка на сайте,запуская часть кода заново, Поехали СЗТ")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            habarovsk1(),
            habarovsk2(),
            habarovsk3(),
            habarovsk4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())
            
def city_krasn():

    async def krasn1():
        while True:
            try:
                useragent = UserAgent()
            #options
                options = webdriver.ChromeOptions()
            #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
            # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
            #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
            ### Омега Красноярск
                driver.get("https://taxiomega.ru/?town=krasnoyarsk")
                await asyncio.sleep(5)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Михаила Годенко7")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Новая Заря, 21","Забобонова, 8", "Ленина168",
                           "Попова, 16","Карла Маркса 98","Брянская19","Ленина20",
                           "Мужества31","Березина 8","Батурина 1"]
                spisok4 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok4
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Красноярск"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка на сайте, перезапуск...Омега")
                driver.quit()
                continue
            driver.quit()
            break

#####Яндекс Красноярск
    async def krasn2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Красноярск Михаила Годенко7")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Новая Заря, 21","Забобонова, 8", "Ленина168",
                           "Попова, 16","Карла Маркса 98","Брянская19","Ленина20",
                           "Мужества31","Березина 8","Батурина 1"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Красноярск"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
###Максим Красноярск
    async def krasn3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=9-Красноярск&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Михаила Годенко7")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Новая Заря, 21","Забобонова, 8", "Ленина168",
                           "Попова, 16","Карла Маркса 98","Брянская19","Ленина20",
                           "Мужества31","Березина 8","Батурина 1"]
                spisok5 = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    driver.implicitly_wait(5)
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok5.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok5
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Красноярск"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка на сайте Максим, перезапуск...")
                driver.quit()
                continue
            driver.quit()
            break

                ###Поехали Красноярск
    async def krasn4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=710-Красноярск&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(2)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Михаила Годенко7")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Новая Заря, 21","Забобонова, 8", "Ленина168",
                           "Попова, 16","Карла Маркса 98","Брянская19","Ленина20",
                           "Мужества31","Березина 8","Батурина 1"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Красноярск"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка на сайте Поехали, перезапуск...")
                driver.quit()
                continue
            driver.quit()
            break

                ####Микс Красноярск
    async def krasn5():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://таксимикс.рф/city/krasnoyarsk.html")
                await asyncio.sleep(4)
                driver.get("http://таксимикс.рф/city/krasnoyarsk.html")
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Михаила Годенко7")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Новая Заря, 21","Забобонова, 8", "Ленина168",
                           "Попова, 16","Карла Маркса 98","Брянская19","Ленина20",
                           "Мужества31","Березина 8","Батурина 1"]
                spisok2 = ["Микс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Красноярск"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка на сайте Микс, перезапуск...")
                driver.quit()
                continue
            driver.quit()
            break

                #### Светофор Красноярск
    async def krasn6():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://svetofor-taxsee.ru")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Михаила Годенко7")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Новая Заря, 21","Забобонова, 8", "Ленина168",
                           "Попова, 16","Карла Маркса 98","Брянская19","Ленина20",
                           "Мужества31","Березина 8","Батурина 1"]
                spisok2 = ["Светофор"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                print(*d)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Красноярск"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка на сайте Светофор, перезапуск...")
                driver.quit()
                pass
            driver.quit()
            break
    async def main():
        tasks = [
            krasn1(),
            krasn2(),
            krasn3(),
            krasn4(),
            krasn5(),
            krasn6()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_chita():
    async def chita1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                #отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                ######Омега Чита
                driver.get("https://taxiomega.ru/?town=chita")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Бутина 37")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Чкалова 141","Нечаева улица, 29","Нечаева улица, 58","Каштакская 7-я улица, 16",
                           "Яблоновая 30","Красной Звезды улица, 76А","Рахова улица, 96",
                           "Рахова улица, 176","Девичья Сопка микрорайон, 29","6 микрорайон, 3"]
                spisok4 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok4
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=240)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Чита"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка сайта Омеги, начинаю проверять")
                driver.quit()
                continue
            driver.quit()
            break
#####Yndex Chita
    async def chita2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Чита Бутина 37")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Чкалова 141","Нечаева улица, 29","Нечаева улица, 58","Каштакская 7-я улица, 16",
                           "Яблоновая 30","Красной Звезды улица, 76А","Рахова улица, 96",
                           "Рахова улица, 176","Девичья Сопка микрорайон, 29","6 микрорайон, 3"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=240)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Чита"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break

                ####Максим Чита
    async def chita3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=59-Чита&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Бутина 37")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Чкалова 141","Нечаева улица, 29","Нечаева улица, 58","Каштакская 7-я улица, 16",
                           "Яблоновая 30","Красной Звезды улица, 76А","Рахова улица, 96",
                           "Рахова улица, 176","Девичья Сопка микрорайон, 29","6 микрорайон, 3"]
                spisok4 = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok4.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    elem1.send_keys(Keys.DELETE)
                    await asyncio.sleep(4)
                    i += 1
                d = spisok4
                if d[1]=="":
                    continue
                t = datetime.now() + timedelta(minutes=240)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Чита"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
                # print(*d)
            except Exception:
                print("Ошибка Максим, начинаю заново")
                driver.quit()
                continue
            driver.quit()
            break
                ####Поехали Чита
    async def chita4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=644-Чита&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                time.sleep(3)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Бутина 37")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Чкалова 141","Нечаева улица, 29","Нечаева улица, 58","Каштакская 7-я улица, 16",
                           "Яблоновая 30","Красной Звезды улица, 76А","Рахова улица, 96",
                           "Рахова улица, 176","Девичья Сопка микрорайон, 29","6 микрорайон, 3"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                print(*d)
                t = datetime.now() + timedelta(minutes=240)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Чита"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка сайта Поехали , начинаю заново")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            chita1(),
            chita2(),
            chita3(),
            chita4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_yroslavl():

    async def yroslavl1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=21-Ярославль&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Ленинградский проспект, 46")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Дзержинского проспект, 6", "Панина 10", "Бабича 5", "Тутаевское шоссе 105",
                            "Промышленное шоссе, 12", "Куропаткова улица 12 А", "Красноперевальская 2-я улица 36", "Михайловская 40",
                            "Республиканская улица 37","Большая Октябрьская 39"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                print(*b)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Ярославль"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка сайта Maxsim, запускаю снова")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Ярославль
    async def yroslavl2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Ярославль Ленинградский проспект, 46")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Дзержинского проспект, 6", "Панина 10", "Бабича 5", "Тутаевское шоссе 105",
                            "Промышленное шоссе, 12", "Куропаткова улица 12 А", "Красноперевальская 2-я улица 36", "Михайловская 40",
                            "Республиканская улица 37","Большая Октябрьская 39"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Ярославль"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
                #Омега сайт Ярославль
    async def yroslavl3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=yaroslavl")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Ленинградский проспект, 46")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Дзержинского проспект, 6", "Панина 10", "Бабича 5", "Тутаевское шоссе 105",
                            "Промышленное шоссе, 12", "Куропаткова улица 12 А", "Красноперевальская 2-я улица 36", "Михайловская 40",
                            "Республиканская улица 37","Большая Октябрьская 39"]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Ярославль"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка сайта omega, запускаю снова")
                driver.quit()
                continue
            driver.quit()
            break
                #Поехали такси Ярославль
    async def yroslavl4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=615-Ярославль&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                time.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Ленинградский проспект, 46")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Дзержинского проспект, 6", "Панина 10", "Бабича 5", "Тутаевское шоссе 105",
                            "Промышленное шоссе, 12", "Куропаткова улица 12 А", "Красноперевальская 2-я улица 36", "Михайловская 40",
                            "Республиканская улица 37","Большая Октябрьская 39"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                print(*b)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Ярославль"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка сайта Poehali, запускаю снова")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            yroslavl1(),
            yroslavl2(),
            yroslavl3(),
            yroslavl4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_kurgan():
    async def kgn():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                #Максим Курган
                driver.get("https://taximaxim.com/ru/?city=1-Курган&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Куйбышева улица, 12")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Красина 66", "Красина 92", "Гоголя 98А","Коли Мяготина улица, 41",
                            "Бажова 12", "Конституции проспект, 69", "Конституции проспект, 41",
                               "Конституции проспект, 26","Конституции проспект, 25","Конституции проспект, 3А"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(3)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    #price1.click()
                    driver.implicitly_wait(5)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Курган"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка, долго прогружаются данные на сайте Максим, начну завново")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Курган
    async def kgn1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Курган Куйбышева улица, 12")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Красина 66", "Красина 94", "Гоголя 98А","Коли Мяготина улица, 41",
                            "Бажова 12", "Конституции проспект, 69", "Конституции проспект, 41",
                               "Конституции проспект, 26","Конституции проспект, 25","Конституции проспект, 3А"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Курган"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break

#Омега Курган
    async def kgn2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                                        #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                                        # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                    #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=default")
                await asyncio.sleep(5)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[1]/div/div[2]/div/div[1]/div/div/span/input[2]")
                await asyncio.sleep(4)
                elem.send_keys("Куйбышева улица, 12")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Красина 66", "Красина 92", "Гоголя 98А","Коли Мяготина улица, 41",
                            "Бажова 12", "Конституции проспект, 69", "Конституции проспект, 41",
                               "Конституции проспект, 26","Конституции проспект, 25","Конституции проспект, 3А"]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                c = spisok1
                print(*c)
                if c[1] == "":
                    continue
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Курган"]
                row = (c + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка, долго прогружаются данные на сайте Омега, начну завново")
                driver.quit()
                continue
            driver.quit()
            break

#Поехали Курган
    async def kgn3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=586-Курган&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(2)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Куйбышева улица, 12")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Красина 66", "Красина 92", "Гоголя 98А","Коли Мяготина улица, 41",
                            "Бажова 12", "Конституции проспект, 69", "Конституции проспект, 41",
                               "Конституции проспект, 26","Конституции проспект, 25","Конституции проспект, 3А"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(3)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                #print(*d)
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Курган"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка, долго прогружаются данные на сайте Поехали, начну завново")
                driver.quit()
                continue
            driver.quit()
            break
####Микс Курган
    async def kgn4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://таксимикс.рф")
                await asyncio.sleep(4)
                driver.get("http://таксимикс.рф")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Куйбышева улица, 12")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Красина 66", "Красина 92", "Гоголя 98А","Коли Мяготина улица, 41",
                            "Бажова 12", "Конституции проспект, 69", "Конституции проспект, 41",
                               "Конституции проспект, 26","Конституции проспект, 25","Конституции проспект, 3А"]
                spisok2 = ["Микс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=0)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Курган"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Ошибка, долго прогружаются данные на сайте Mix, начну заново")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            kgn(),
            kgn1(),
            kgn2(),
            kgn3(),
            kgn4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_ykutsk():
    async def city1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
#Максим Якутск
                driver.get("https://taximaxim.com/ru/?city=272-Якутск&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Авиаторов улица, 15")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Авиаторов улица, 66","Намская улица, 1","Автострада 50 лет Октября, 7","Кычкина 4",
                           "Дзержинского, 53","Кальвица, 18","Дзержинского улица, 14","Курашова улица, 8",
                           "Ярославского улица, 28","Дежнева улица, 17"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=240)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Якутск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Максим, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Якутск
    async def city2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Якутск Авиаторов улица, 15")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Авиаторов улица, 66","Намская улица, 1","Автострада 50 лет Октября, 7","Кычкина 4",
                           "Дзержинского, 53","Кальвица, 18","Дзержинского улица, 14","Курашова улица, 8",
                           "Ярославского улица, 28","Дежнева улица, 17"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=240)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Якутск"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break

                #Омега сайт Якутск
    async def city3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=yakutsk")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Авиаторов улица, 15")
                await asyncio.sleep(5)
                time.sleep(3)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Авиаторов улица, 66","Намская улица, 1","Автострада 50 лет Октября, 7","Кычкина 4",
                           "Дзержинского, 53","Кальвица, 18","Дзержинского улица, 14","Курашова улица, 8",
                           "Ярославского улица, 28","Дежнева улица, 17"]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                if b[1]=="":
                    continue
                t = datetime.now() + timedelta(minutes=240)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Якутск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
                #print(*b)
            except Exception:
                print("Произошла ошибка на сайте Омега, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break

                #Поехали Якутск
    async def city4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=817-Якутск&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Авиаторов улица, 15")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Авиаторов улица, 66","Намская улица, 1","Автострада 50 лет Октября, 7","Кычкина 4",
                           "Дзержинского, 53","Кальвица, 18","Дзержинского улица, 14","Курашова улица, 8",
                           "Ярославского улица, 28","Дежнева улица, 17"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                print(*b)
                t = datetime.now() + timedelta(minutes=240)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Якутск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Поехали, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            city1(),
            city2(),
            city3(),
            city4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_vladik():
    async def v1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=25-Владивосток&referrer=self")
                await asyncio.sleep(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Океанский проспект, 100А")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Некрасовская 100","Народный проспект 17","Красного знамени 156",
                           "Верхнепортовая улица, 3","Верхнепортовая 32","Нижнепортовая 6г","Крыгина 19а",
                           "Леонова 66","Зои Космодемьянской 40","Сахалинская 25"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Владивосток"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Максим, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Владивосток
    async def v2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Владивосток Океанский проспект, 100А")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Некрасовская 100","Народный проспект 17","Красного знамени 156",
                           "Верхнепортовая улица, 3","Верхнепортовая 32","Нижнепортовая 6г","Крыгина 19а",
                           "Леонова 66","Зои Космодемьянской 40","Сахалинская 25"]
                spisok = ["Яндекс:  "]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Владивосток"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
                #Омега сайт Владивосток
    async def v3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=vladivostok")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Океанский проспект, 100А")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Некрасовская 100","Народный проспект 17","Красного знамени 156",
                           "Верхнепортовая улица, 3","Верхнепортовая 32","Нижнепортовая 6г","Крыгина 19а",
                           "Леонова 66","Зои Космодемьянской 40","Сахалинская 25"]
                spisok1 = ["Омега:    "]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Владивосток"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Омега, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
                #Поехали Владивосток
    async def v4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=637-Владивосток&intl=ru-RU&referrer=self")
                await asyncio.sleep(5)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                await asyncio.sleep(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Океанский проспект, 100А")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Некрасовская 100","Народный проспект 17","Красного знамени 156",
                           "Верхнепортовая улица, 3","Верхнепортовая 32","Нижнепортовая 6г","Крыгина 19а",
                           "Леонова 66","Зои Космодемьянской 40","Сахалинская 25"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                print(*b)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Владивосток"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Поехали, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
####Микс Владивосток
    async def v5():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://таксимикс.рф/city/vladivostok.html")
                await asyncio.sleep(4)
                driver.get("http://таксимикс.рф/city/vladivostok.html")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Океанский проспект, 100А")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Некрасовская 100","Народный проспект 17","Красного знамени 156",
                           "Верхнепортовая улица, 3","Верхнепортовая 32","Нижнепортовая 6г","Крыгина 19а",
                           "Леонова 66","Зои Космодемьянской 40","Сахалинская 25"]
                spisok2 = ["Микс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                print(*d)
                t = datetime.now() + timedelta(minutes=300)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Владивосток"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Микс, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            v1(),
            v2(),
            v3(),
            v4(),
            v5()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_barnaul():
    async def bar1():
        while True:
            try:
                t = datetime.today()
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=11-Барнаул&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Балтийская2")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Балтийская 34", "Балтийская улица, 88", "Сергея Ускова улица, 23",
                           "Попова улица, 242", "Попова55","Трактовая улица, 41/3","Новосибирская улица, 16В",
                           "Новосибирская улица, 46","Баумана улица, 43", "Красносельская50"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Барнаул"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Максим, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Барнаул
    async def bar2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Барнаул Балтийская2")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Балтийская 34", "Балтийская улица, 88", "Сергея Ускова улица, 23",
                           "Попова улица, 242", "Попова55","Трактовая улица, 41/3","Новосибирская улица, 16В",
                           "Новосибирская улица, 46","Баумана улица, 43", "Красносельская50"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Барнаул"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
                #Омега сайт Барнаул
    async def bar3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=barnaul")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Балтийская2")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Балтийская 34", "Балтийская улица, 88", "Сергея Ускова улица, 23",
                           "Попова улица, 242", "Попова55","Трактовая улица, 41/3","Новосибирская улица, 16В",
                           "Новосибирская улица, 46","Баумана улица, 43", "Красносельская50"]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                if b[1]=="":
                    continue
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Барнаул"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
                #print(*b)
            except Exception:
                print("Произошла ошибка на сайте Омега, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break

                #Поехали Барнаул
    async def bar4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=654-Барнаул&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Балтийская2")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Балтийская 34", "Балтийская улица, 88", "Сергея Ускова улица, 23",
                           "Попова улица, 242", "Попова55","Трактовая улица, 41/3","Новосибирская улица, 16В",
                           "Новосибирская улица, 46","Баумана улица, 43", "Красносельская50"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                print(*b)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Барнаул"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Поехали, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            bar1(),
            bar2(),
            bar3(),
            bar4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_irkutsk():
    async def ir1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=7-Иркутск&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Тимирязева, 23Б")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Байкальская улица, 10","Байкальская улица, 50","Депутатская улица, 48",
                           "Байкальская улица, 157/1", "Байкальская улица, 250Б/9",
                           "Байкальская улица, 239","Маршала Жукова проспект, 15/4","Рабочего Штаба улица, 134/2",
                           "Рябикова бульвар, 32Г/11","Рябикова бульвар, 47"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Иркутск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
                #print(*b)
            except Exception:
                print("Произошла ошибка - долгое ожидание от сайта Maxim, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Иркутск
    async def ir2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Иркутск Тимирязева, 23Б")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Байкальская улица, 10","Байкальская улица, 50","Депутатская улица, 48",
                           "Байкальская улица, 157/1", "Байкальская улица, 250Б/9",
                           "Байкальская улица, 239","Маршала Жукова проспект, 15/4","Рабочего Штаба улица, 134/2",
                           "Рябикова бульвар, 32Г/11","Рябикова бульвар, 47"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Иркутск"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break
    async def ir3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                #Омега сайт Иркутск
                driver.get("https://taxiomega.ru/?town=irkutsk")
                await asyncio.sleep(5)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Тимирязева, 23Б")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Байкальская улица, 10","Байкальская улица, 50","Депутатская улица, 48",
                           "Байкальская улица, 157/1", "Байкальская улица, 250Б/9",
                           "Байкальская улица, 239","Маршала Жукова проспект, 15/4","Рабочего Штаба улица, 134/2",
                           "Рябикова бульвар, 32Г/11","Рябикова бульвар, 47"]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                if b[1]=="":
                    continue
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Иркутск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
               # print(*b)
            except Exception:
                print("Произошла ошибка - долгое ожидание от сайта Omega, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break

                #Поехали Иркутск
    async def ir4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=662-Иркутск&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Тимирязева, 23Б")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Байкальская улица, 10","Байкальская улица, 50","Депутатская улица, 48",
                           "Байкальская улица, 157/1", "Байкальская улица, 250Б/9",
                           "Байкальская улица, 239","Маршала Жукова проспект, 15/4","Рабочего Штаба улица, 134/2",
                           "Рябикова бульвар, 32Г/11","Рябикова бульвар, 47"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                t = datetime.now() + timedelta(minutes=180)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Иркутск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка - долгое ожидание от сайта Poexali, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            ir1(),
            ir2(),
            ir3(),
            ir4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

def city_archangelsk():
    async def arh1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=48-Архангельск&referrer=self")
                await asyncio.sleep(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Чкалова, 5")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Луговая улица, 15","Гоголя 6","Первомайская 17",
                           "Ленинградский проспект, 15/1","Урицкого, 33","Розы Люксембург улица, 57",
                           "Троицкий проспект, 62","Троицкий проспект, 81","Троицкий проспект, 123","Талажское шоссе, 22"
                           ]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Архангельск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
                #print(*b)
            except Exception as ex:
                print("Ошибка ожидания от сайта Maxim")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Архангельск
    async def arh2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Архангельск Чкалова, 5")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["Луговая улица, 15","Гоголя 6","Первомайская 17",
                           "Ленинградский проспект, 15/1","Урицкого, 33","Розы Люксембург улица, 57",
                           "Троицкий проспект, 62","Троицкий проспект, 81","Троицкий проспект, 123","Талажское шоссе, 22"
                           ]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Архангельск"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break

                #Омега сайт Архангельск
    async def arh3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=arhangelsk")
                await asyncio.sleep(5)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Чкалова, 5")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Луговая улица, 15","Гоголя 6","Первомайская 17",
                           "Ленинградский проспект, 15/1","Урицкого, 33","Розы Люксембург улица, 57",
                           "Троицкий проспект, 62","Троицкий проспект, 81","Троицкий проспект, 123","Талажское шоссе, 22"
                           ]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                if b[1] == "":
                    continue
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Архангельск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
                #print(*b)
            except Exception as ex:
                print("Ошибка ожидания от сайта Omega")
                driver.quit()
                continue
            driver.quit()
            break

                #Поехали такси Архангельск
    async def arh4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=887-Архангельск&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(2)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Чкалова улица, 5")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Луговая улица, 15","Гоголя 6","Первомайская 17",
                           "Ленинградский проспект, 15/1","Урицкого, 33","Розы Люксембург улица, 57",
                           "Троицкий проспект, 62","Троицкий проспект, 81","Троицкий проспект, 123","Талажское шоссе, 22"
                           ]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(3)
                    i += 1
                b = spisok2
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Архангельск"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Poexali")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            arh1(),
            arh2(),
            arh3(),
            arh4()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())
            
def city_nabcheln():
    async def nh1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
        ###НЧ
                driver.get("https://taximaxim.com/ru/?city=84-Набережные+Челны&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Мира проспект, 3")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["31-й комплекс, 17А","59-й комплекс, 15","56-й комплекс, 18","43-й комплекс, 6", "45-й комплекс, 8",
                           "9-й комплекс (старый Город), 2","26-й комплекс, 23","5-й комплекс (Старый Город), 6",
                           "7-й комплекс (Старый Город), 22","Казанский проспект, 15С/31"]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                if b[1]=="":
                    continue
                print(*b)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Набережные"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Maxim, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
#####Яндекс Набережные челны
    async def nh2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://passport.yandex.ru/auth/add?retpath=https%3A%2F%2Ftaxi.yandex.ru%2F&from=taxi")
                await asyncio.sleep(4)
                tel = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[1]/span/input")
                tel.send_keys("#####")#<--логин
                driver.implicitly_wait(5)
                open1 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                log = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div/form/div[2]/span/input")
                log.send_keys("######")#<--Пароль
                driver.implicitly_wait(5)
                open2 = driver.find_element_by_xpath("/html/body/div/div/div[2]/div[2]/div/div/div[2]/div[3]/div/div/div[1]/form/div[3]/button").click()
                driver.implicitly_wait(5)
                string1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]").click()
                await asyncio.sleep(4)
                krest1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[1]").click()
                time.sleep(3)
                #driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[1]/div[2]/span[1]/span[2]/textarea")
                elem.send_keys("Набережные Челны Мира проспект, 3")
                await asyncio.sleep(4)
                vibor1 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                elem1 = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[2]/textarea")
                address = ["31-й комплекс, 17А","59-й комплекс, 15","56-й комплекс, 18","43-й комплекс, 6", "45-й комплекс, 8",
                           "Камснаб Торговый комплекс","26-й комплекс, 23","5-й комплекс, 6",
                           "7-й комплекс, 22","Казанский проспект, 15С/31"]
                spisok = ["Яндекс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem22 = driver.find_element_by_class_name("_1raGI4VPNMAm6ssRfDpaIm").click()
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_class_name("_1_gHsnrkwYevH989PurjaV")
                    #"/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div[2]/div[1]/div/div[2]/div/div/div/button[1]/div[2]/div[3]/span[2]")
                    await asyncio.sleep(4)
                    a = price1.text.replace("\u202f₽", " ")
                    spisok.append(a)
                    elem1.click()
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[1]/div[2]/div[1]/div[4]/div/div[1]/div/div[1]/div/div[1]/div[2]/div[2]/span[1]/span[1]").click()
                    await asyncio.sleep(4)
                    i += 1
                a = spisok
                print(*a)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Набережные"]
                row = (a + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Яндекса")
                driver.quit()
                #pass
                continue
            driver.quit()
            break

                #Омега сайт НЧ
    async def nh3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=naberejnye_chelny")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                time.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Мира проспект, 3")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["31-й комплекс, 17А","59-й комплекс, 15","56-й комплекс, 18","43-й комплекс, 6", "45-й комплекс, 8",
                           "9-й комплекс (старый Город), 2","26-й комплекс, 23","5-й комплекс (Старый Город), 6",
                           "7-й комплекс (Старый Город), 22","Казанский проспект, 15С/31"]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                if b[1] == "":
                    continue
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Набережные"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
                #print(*b)
            except Exception:
                print("Произошла ошибка на сайте Омега, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break

                #Поехали НЧ
    async def nh4():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=712-Набережные+Челны&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Мира проспект, 3")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["31-й комплекс, 17А","59-й комплекс, 15","56-й комплекс, 18","43-й комплекс, 6", "45-й комплекс, 8",
                           "9-й комплекс (старый Город), 2","26-й комплекс, 23","5-й комплекс (Старый Город), 6",
                           "7-й комплекс (Старый Город), 22","Казанский проспект, 15С/31"]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    driver.implicitly_wait(5)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                if b[1]=="":
                    continue
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Набережные"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
               # print(*b)
            except Exception:
                print("Произошла ошибка на сайте Поехали, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
####Микс НЧ МИкс
    async def nh5():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("http://таксимикс.рф/city/nchelnyn.html")
                await asyncio.sleep(4)
                driver.get("http://таксимикс.рф/city/nchelnyn.html")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Мира проспект, 3")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["31-й комплекс, 17А","59-й комплекс, 15","56-й комплекс, 18","43-й комплекс, 6", "45-й комплекс, 8",
                           "9-й комплекс (старый Город), 2","26-й комплекс, 23","5-й комплекс (Старый Город), 6",
                           "7-й комплекс (Старый Город), 22","Казанский проспект, 15С/31"]
                spisok2 = ["Микс"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[3]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i").click()
                    await asyncio.sleep(4)
                    i += 1
                d = spisok2
                if d[1]=="":
                    continue
                print(*d)
                t = datetime.now() + timedelta(minutes=-120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Набережные"]
                row = (d + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception:
                print("Произошла ошибка на сайте Микс, начну сначала")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            nh1(),
            nh2(),
            nh3(),
            nh4(),
            nh5()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())
    ####Новокузнецк
def city_novokuzn():
    async def nk1():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taximaxim.com/ru/?city=53-Новокузнецк&referrer=self")
                await asyncio.sleep(4)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Веры Соломиной, 36")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Карла Маркса, 10","Челюскина, 7","Рудокопровая, 33","Куйбышева, 18",
                           "Курако проспект, 45","Энтузиастов, 30","Кирова 57",
                            "Кирова 70","Кирова 96","Транспортная, 127"
                           ]
                spisok = ["Максим"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok
                print(*b)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Новокузнецк"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Maxim")
                driver.quit()
                continue
            driver.quit()
            break

                #Омега сайт Новокузнецк
    async def nk2():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxiomega.ru/?town=novokuzneck")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/section[1]/div[2]/div/div[1]/a").click()
                driver.implicitly_wait(5)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Веры Соломиной, 36")
                await asyncio.sleep(5)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Карла Маркса, 10","Челюскина, 7","Рудокопровая, 33","Куйбышева, 18",
                           "Курако проспект, 45","Энтузиастов, 30","Кирова 57",
                            "Кирова 70","Кирова 96","Транспортная, 127"
                           ]
                spisok1 = ["Омега"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[6]/div/div[1]/div/span")
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok1.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div[2]/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok1
                print(*b)
                if b[1] == "":
                    continue
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Новокузнецк"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Omega")
                driver.quit()
                continue
            driver.quit()
            break

                #Поехали такси Новокузнецк
    async def nk3():
        while True:
            try:
                useragent = UserAgent()
                #options
                options = webdriver.ChromeOptions()
                #options.add_argument("user-agent=HelloWorld:)")
                options.add_argument(f"user-agent={useragent.random}") #Рандомный useragent
                # отключение автоматизации(чтобы сайт не распознавал автоматизацию)
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_argument("headless") #включение фонового режима
                driver = webdriver.Chrome(
                #executable_path = "D:\\Прога разработка\\parsing\\chromedriver.exe",   # путь
                    options=options)
                driver.get("https://taxipoehali.ru/client/?city=603-Новокузнецк&intl=ru-RU&referrer=self")
                await asyncio.sleep(4)
                order = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[2]/button[2]").click()
                driver.implicitly_wait(2)
                driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
                elem = driver.find_element_by_id("addressform-0-pointfield")
                elem.send_keys("Веры Соломиной, 36")
                await asyncio.sleep(4)
                elem.send_keys(Keys.TAB)
                elem1 = driver.find_element_by_id("addressform-1-pointfield")
                address = ["Карла Маркса, 10","Челюскина, 7","Рудокопровая, 33","Куйбышева, 18",
                           "Курако проспект, 45","Энтузиастов, 30","Кирова 57",
                            "Кирова 70","Кирова 96","Транспортная, 127"
                           ]
                spisok2 = ["Поехали"]
                i = 0
                while i < len(address):
                    elem1.send_keys(address[i])
                    await asyncio.sleep(4)
                    elem1.send_keys(Keys.TAB)
                    await asyncio.sleep(4)
                    price1 = driver.find_element_by_id("price")
                    #price1.click()
                    await asyncio.sleep(4)
                    a = price1.text.replace("₽", "")
                    spisok2.append(a)
                    cross = driver.find_element_by_xpath("/html/body/div[1]/div/div[2]/div[1]/form/div[1]/div[1]/div/div[2]/div/div[2]/div/div[1]/div/div/div/div/button/i")
                    cross.click()
                    await asyncio.sleep(4)
                    i += 1
                b = spisok2
                print(*b)
                t = datetime.now() + timedelta(minutes=120)
                am = [t]
                fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"
                wb = load_workbook(fn)
                ws = wb["Новокузнецк"]
                row = (b + am)   # <--- новая строка
                ws.append(row)
                wb.save(fn)
                wb.close()
                driver.quit()
            except Exception as ex:
                print("Ошибка ожидания от сайта Poexali")
                driver.quit()
                continue
            driver.quit()
            break
    async def main():
        tasks = [
            nk1(),
            nk2(),
            nk3()
        ]
        await asyncio.gather(*tasks)
    asyncio.run(main())

abcde = input("Введите название города(ов)ЧЕРЕЗ ЗАПЯТУЮ:\n(Тула,Казань,Улан,Томск,Тюмень,Кемерово,Сургут,\nХабаровск,Красноярск,Чита,Ярославль,Курган,Якутск,Владивосток,Барнаул,Иркутск,Архангельск,Набережные,Новокузнецк,): ")
time_a = int(input("Введите время ожидания между сбором данных в МИНУТАХ:"))
time_b = time_a*60
spisok = ["***","***","***","***","***","***","***","***","***","***","***","***"]
bb = spisok
while True:
    if "Тула" in abcde:
        city_tula()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Тула"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Казань" in abcde:
        city_kazan()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Казань"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Улан" in abcde:
        city_ulan()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Улан"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Томск" in abcde:
        city_tomsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Томск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Тюмень" in abcde:
        city_tymen()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Тюмень"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Кемерово" in abcde:
        city_kemerovo()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Кемерово"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Сургут" in abcde:
        city_surgut()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Сургут"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Хабаровск" in abcde:
        city_habarovsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Хабаровск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Красноярск" in abcde:
        city_krasn()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Красноярск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Чита" in abcde:
        city_chita()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Чита"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Ярославль" in abcde:
        city_yroslavl()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Ярославль"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Курган" in abcde:
        city_kurgan()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Курган"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Якутск" in abcde:
        city_ykutsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Якутск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Владивосток" in abcde:
        city_vladik()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Владивосток"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Барнаул" in abcde:
        city_barnaul()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Барнаул"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Иркутск" in abcde:
        city_irkutsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Иркутск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Архангельск" in abcde:
        city_archangelsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Архангельск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Набережные" in abcde:
        city_nabcheln()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Набережные"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
    if "Новокузнецк" in abcde:
        city_novokuzn()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Новокузнецк"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        time.sleep(time_b)
        
    if "Все" in abcde:
        city_tula()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Тула"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        #time.sleep(time_b)
        city_kazan()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Казань"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
    
        city_ulan()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Улан"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_tomsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Томск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_tymen()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Тюмень"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_kemerovo()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Кемерово"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_surgut()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Сургут"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_habarovsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Хабаровск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_krasn()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Красноярск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_chita()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Чита"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_yroslavl()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Ярославль"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_kurgan()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Курган"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_ykutsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Якутск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_vladik()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Владивосток"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_barnaul()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Барнаул"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_irkutsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Иркутск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_archangelsk()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Архангельск"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_nabcheln()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Набережные"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()

        city_novokuzn()
        fn = r"C:\Users\Илья_К\Desktop\Мониторинг_все_города/city_full.xlsx"   
        wb = load_workbook(fn)    
        ws = wb["Новокузнецк"]   
        row = (bb)   # <--- новая строка    
        ws.append(row)    
        wb.save(fn)    
        wb.close()
        
        
    
    

    
